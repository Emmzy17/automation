import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.chart import BarChart, Reference
wb = openpyxl.load_workbook('sales.xlsx')
sheets = wb.sheetnames
userSheet = wb['Users']
saleSheet = wb['Sales']
#newSheet = wb.create_sheet('new sheet')
#print(wb.active)
#wb.save('sales.xlsx')
#wb.remove_sheet('new sheet1')
#wb.save('sales.xlsx')
#dict_cell = userSheet._cells
#for column in userSheet.columns:
#    print(column[0].value,column[1].value, column[2].value,column[3].value)
#print(userSheet['c6'].value)
#userSheet['e1'] = 'New Total'
ref = Reference(userSheet, min_col=3, min_row=2, max_col=3, max_row=11)
chart = BarChart()
chart.add_data(ref)
userSheet.add_chart(chart, 'J6')

for i in range(2, 12):
    userSheet['E' + str(i)] = userSheet['C' + str(i)].value + 5
userSheet['E12'] = '=SUM(E1:E11)'
img= Image('new-logo-csk-2.png')
img2 = Image('new-logo-csk-2.png')

#saleSheet.merge_cells('C2:D3')



#userSheet.add_image(img, 'B12')
#userSheet.add_image(img2, 'E12')
wb.save('sales.xlsx')


print(saleSheet.dimensions)
print(userSheet['b2'].value)

